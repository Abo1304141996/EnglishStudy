"""
extract_xlsx.py
提取 shameless02 目录下所有 .xlsx 文件的内容，输出为 CSV / JSON / 终端预览
用法:
    python scripts/extract_xlsx.py                        # 预览所有 xlsx
    python scripts/extract_xlsx.py --output csv           # 每个 sheet 输出为 csv
    python scripts/extract_xlsx.py --output json          # 每个 sheet 输出为 json
    python scripts/extract_xlsx.py --file shameless02/Shameless-S02E02.xlsx
"""

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent.parent  # 项目根目录


def load_sheets(xlsx_path: Path) -> dict[str, list[list]]:
    """读取 xlsx 文件，返回 {sheet_name: [[row], ...]}"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(list(row))
        result[name] = rows
    return result


def preview(path: Path, sheets: dict):
    print(f"\n{'='*60}")
    print(f"文件: {path.name}")
    for sheet_name, rows in sheets.items():
        print(f"\n  [Sheet: {sheet_name}]  共 {len(rows)} 行")
        for i, row in enumerate(rows[:10]):
            print(f"    {row}")
        if len(rows) > 10:
            print(f"    ... (共 {len(rows)} 行，仅显示前 10 行)")


def export_csv(path: Path, sheets: dict, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    for sheet_name, rows in sheets.items():
        safe_sheet = sheet_name.replace("/", "_").replace("\\", "_")
        out_file = out_dir / f"{path.stem}__{safe_sheet}.csv"
        with open(out_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        print(f"  ✅ CSV 已写入: {out_file}")


def export_json(path: Path, sheets: dict, out_dir: Path):
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{path.stem}.json"
    data = {sheet: rows for sheet, rows in sheets.items()}
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  ✅ JSON 已写入: {out_file}")


def main():
    parser = argparse.ArgumentParser(description="提取 xlsx 数据")
    parser.add_argument("--file", help="指定单个 xlsx 文件路径（相对项目根目录）")
    parser.add_argument("--output", choices=["preview", "csv", "json"], default="preview")
    parser.add_argument("--out-dir", default="scripts/output", help="输出目录（默认 scripts/output）")
    args = parser.parse_args()

    if args.file:
        xlsx_files = [ROOT / args.file]
    else:
        # 默认扫描 shameless02 目录
        xlsx_files = list((ROOT / "shameless02").glob("*.xlsx"))

    if not xlsx_files:
        print("未找到任何 xlsx 文件")
        sys.exit(1)

    out_dir = ROOT / args.out_dir

    for xlsx_path in sorted(xlsx_files):
        if not xlsx_path.exists():
            print(f"文件不存在: {xlsx_path}")
            continue
        print(f"\n处理: {xlsx_path}")
        sheets = load_sheets(xlsx_path)

        if args.output == "csv":
            export_csv(xlsx_path, sheets, out_dir)
        elif args.output == "json":
            export_json(xlsx_path, sheets, out_dir)
        else:
            preview(xlsx_path, sheets)


if __name__ == "__main__":
    main()
