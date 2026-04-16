"""
extract_dialogue.py
从整理后的 xlsx 中提取对话数据（英文-中文对）

规则：
- 该行恰好有 2 列非空 → 对话行，提取为 英文 | 中文
- 该行只有 1 列非空   → 词汇解释行，跳过
- 其他（0列/标题）    → 跳过

用法：
    python scripts/extract_dialogue.py --file shameless02/Shameless-S02E01.xlsx
    python scripts/extract_dialogue.py --file shameless02/Shameless-S02E01.xlsx --output txt
"""

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent.parent


def clean_cell(val) -> str:
    """清理单元格：去除时间戳 [MM:SS]、首尾空白、换行"""
    if val is None:
        return ""
    text = str(val).strip()
    # 去除时间戳 如 [02:32] 或 [1:02:32]
    text = re.sub(r'\[\d{1,2}:\d{2}(?::\d{2})?\]', '', text).strip()
    # 合并内部换行为空格
    text = re.sub(r'\s*\n\s*', ' ', text).strip()
    return text


def is_english(text: str) -> bool:
    """判断是否主要为英文（ASCII 字符占比 > 70%）"""
    if not text:
        return False
    ascii_count = sum(1 for c in text if ord(c) < 128)
    return ascii_count / len(text) > 0.7


def extract_dialogues(xlsx_path: Path) -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    dialogues = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            # 过滤非空单元格
            non_empty = [clean_cell(cell) for cell in row if clean_cell(cell)]

            # 只处理恰好 2 列非空的行
            if len(non_empty) != 2:
                continue

            a, b = non_empty

            # 判断哪个是英文哪个是中文
            if is_english(a) and not is_english(b):
                english, chinese = a, b
            elif is_english(b) and not is_english(a):
                english, chinese = b, a
            else:
                # 两个都是英文或都是中文，跳过
                continue

            # 过滤掉太短的（标题/单字）
            if len(english) < 2 or len(chinese) < 1:
                continue

            dialogues.append((english, chinese))

    return dialogues


def main():
    parser = argparse.ArgumentParser(description="提取 xlsx 对话数据")
    parser.add_argument("--file", required=True, help="xlsx 文件路径（相对项目根目录）")
    parser.add_argument("--output", choices=["print", "txt"], default="print")
    parser.add_argument("--out-dir", default="scripts/output")
    args = parser.parse_args()

    xlsx_path = ROOT / args.file
    if not xlsx_path.exists():
        print(f"文件不存在: {xlsx_path}")
        sys.exit(1)

    print(f"处理: {xlsx_path.name} ...")
    dialogues = extract_dialogues(xlsx_path)
    print(f"提取到 {len(dialogues)} 条对话\n")

    lines = [f"{en}\t{zh}" for en, zh in dialogues]

    if args.output == "txt":
        out_dir = ROOT / args.out_dir
        out_dir.mkdir(parents=True, exist_ok=True)
        out_file = out_dir / f"{xlsx_path.stem}_dialogues.txt"
        with open(out_file, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        print(f"✅ 已写入: {out_file}")
    else:
        for line in lines:
            print(line)


if __name__ == "__main__":
    main()
